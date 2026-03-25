"""
excel_loader.py — Leitura de planilhas (.xlsx, .xls, .csv).
Suporta:
  - Planilha EAN  → {sku: {"ean": "...", "descricao": "..."}}
  - Planilha XML F5 → {sku: {"siscomex": Decimal, "afrmm": Decimal}}
"""

import io
import csv
import zipfile as zfcheck
from typing import Dict, List, Optional, Tuple

from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _limpar_digitos(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def _dec(val) -> Decimal:
    if val is None:
        return Decimal("0")
    t = str(val).strip().replace(",", ".")
    if t.endswith(".0"):
        t = t[:-2]
    try:
        return Decimal(t)
    except InvalidOperation:
        return Decimal("0")


def ler_csv_tolerante(data: bytes) -> Tuple[List[List[str]], str]:
    text = data.decode("utf-8", errors="replace").replace("\x00", "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    try:
        dialect = csv.Sniffer().sniff(text[:5000])
        delim = dialect.delimiter
    except Exception:
        delim = ";"
    f = io.StringIO(text, newline="")
    reader = csv.reader(f, delimiter=delim, quotechar='"', doublequote=True)
    return list(reader), delim


def _find_col(header: List[str], nomes: List[str]) -> Optional[int]:
    for idx, h in enumerate(header):
        if h.strip().lower() in nomes:
            return idx + 1  # 1-based (openpyxl)
    return None


# ---------------------------------------------------------------------------
# Planilha EAN/GTIN  →  {sku: {"ean": str, "descricao": str}}
# ---------------------------------------------------------------------------

def carregar_planilha_ean(data: bytes, nome: str) -> Dict[str, Dict]:
    """
    Lê planilha de EANs.
    Tenta detectar colunas por cabeçalho; fallback: C=SKU, D=EAN.
    """
    nome_lower = nome.lower()

    # XLSX
    if zfcheck.is_zipfile(io.BytesIO(data)) or nome_lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip().lower() for c in ws[1]]

        col_sku = (
            _find_col(header, ["sku", "código", "codigo", "código (sku)", "codigo (sku)", "cód", "cod"])
            or 3
        )
        col_ean = (
            _find_col(header, ["ean", "gtin", "código de barras", "codigo de barras", "barcode"])
            or 4
        )
        col_desc = _find_col(header, ["descrição", "descricao", "desc", "produto", "xprod"]) or None

        resultado: Dict[str, Dict] = {}
        for row in range(2, ws.max_row + 1):
            sku = str(ws.cell(row=row, column=col_sku).value or "").strip()
            if not sku:
                continue
            ean_raw = str(ws.cell(row=row, column=col_ean).value or "").strip()
            if ean_raw.endswith(".0"):
                ean_raw = ean_raw[:-2]
            ean = _limpar_digitos(ean_raw)
            desc = ""
            if col_desc:
                desc = str(ws.cell(row=row, column=col_desc).value or "").strip()
            if ean:
                resultado[sku] = {"ean": ean, "descricao": desc}
        return resultado

    # XLS antigo
    if data[:4] == b"\xD0\xCF\x11\xE0" or nome_lower.endswith(".xls"):
        try:
            import xlrd
        except ImportError as e:
            raise ValueError("Arquivo .xls requer: pip install xlrd") from e
        book = xlrd.open_workbook(file_contents=data)
        ws = book.sheet_by_index(0)
        resultado: Dict[str, Dict] = {}
        for r in range(1, ws.nrows):
            sku = str(ws.cell_value(r, 2)).strip()
            if not sku:
                continue
            ean_raw = str(ws.cell_value(r, 3)).strip()
            if ean_raw.endswith(".0"):
                ean_raw = ean_raw[:-2]
            ean = _limpar_digitos(ean_raw)
            if ean:
                resultado[sku] = {"ean": ean, "descricao": ""}
        return resultado

    # CSV
    rows, _ = ler_csv_tolerante(data)
    if not rows:
        return {}
    header = [h.strip().lower() for h in rows[0]]
    sku_i = None
    ean_i = None
    for i, h in enumerate(header):
        if h in ["sku", "código", "codigo", "código (sku)", "codigo (sku)"]:
            sku_i = i
        if h in ["ean", "gtin", "código de barras", "codigo de barras"]:
            ean_i = i
    start = 1 if (sku_i is not None and ean_i is not None) else 0
    sku_i = sku_i if sku_i is not None else 2
    ean_i = ean_i if ean_i is not None else 3

    resultado: Dict[str, Dict] = {}
    for r in rows[start:]:
        if len(r) <= max(sku_i, ean_i):
            continue
        sku = str(r[sku_i]).strip()
        if not sku:
            continue
        ean_raw = str(r[ean_i]).strip()
        if ean_raw.endswith(".0"):
            ean_raw = ean_raw[:-2]
        ean = _limpar_digitos(ean_raw)
        if ean:
            resultado[sku] = {"ean": ean, "descricao": ""}
    return resultado


# ---------------------------------------------------------------------------
# Planilha XML F5  →  {sku: {"siscomex": Decimal, "afrmm": Decimal}}
# ---------------------------------------------------------------------------

def carregar_planilha_xmlf5(data: bytes, nome: str) -> Dict[str, Dict]:
    """
    Lê Planilha XML F5 (Siscomex/AFRMM).
    Tenta detectar colunas por cabeçalho; fallback: C=SKU, D=Siscomex, E=AFRMM.
    """
    nome_lower = nome.lower()

    # XLSX
    if zfcheck.is_zipfile(io.BytesIO(data)) or nome_lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip().lower() for c in ws[1]]

        col_sku = (
            _find_col(header, ["sku", "código", "codigo", "código (sku)", "codigo (sku)"])
            or 3
        )
        col_sis = _find_col(header, ["siscomex"]) or 4
        col_afr = _find_col(header, ["afrmm"]) or 5

        resultado: Dict[str, Dict] = {}
        for row in range(2, ws.max_row + 1):
            sku = str(ws.cell(row=row, column=col_sku).value or "").strip()
            if not sku:
                continue
            sis = _dec(ws.cell(row=row, column=col_sis).value)
            afr = _dec(ws.cell(row=row, column=col_afr).value)
            resultado[sku] = {"siscomex": sis, "afrmm": afr}
        return resultado

    # XLS antigo
    if data[:4] == b"\xD0\xCF\x11\xE0" or nome_lower.endswith(".xls"):
        try:
            import xlrd
        except ImportError as e:
            raise ValueError("Arquivo .xls requer: pip install xlrd") from e
        book = xlrd.open_workbook(file_contents=data)
        ws = book.sheet_by_index(0)
        resultado: Dict[str, Dict] = {}
        for r in range(1, ws.nrows):
            sku = str(ws.cell_value(r, 2)).strip()
            if not sku:
                continue
            resultado[sku] = {
                "siscomex": _dec(ws.cell_value(r, 3)),
                "afrmm": _dec(ws.cell_value(r, 4)),
            }
        return resultado

    # CSV
    rows, _ = ler_csv_tolerante(data)
    if not rows:
        return {}
    header = [h.strip().lower() for h in rows[0]]
    sku_i = None
    sis_i = None
    afr_i = None
    for i, h in enumerate(header):
        if h in ["sku", "código", "codigo", "código (sku)", "codigo (sku)"]:
            sku_i = i
        if h == "siscomex":
            sis_i = i
        if h == "afrmm":
            afr_i = i
    start = 1 if (sku_i is not None) else 0
    sku_i = sku_i if sku_i is not None else 2
    sis_i = sis_i if sis_i is not None else 3
    afr_i = afr_i if afr_i is not None else 4

    resultado: Dict[str, Dict] = {}
    for r in rows[start:]:
        if len(r) <= max(sku_i, sis_i, afr_i):
            continue
        sku = str(r[sku_i]).strip()
        if not sku:
            continue
        resultado[sku] = {"siscomex": _dec(r[sis_i]), "afrmm": _dec(r[afr_i])}
    return resultado
